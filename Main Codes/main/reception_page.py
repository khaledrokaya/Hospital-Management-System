import tkinter as tk
import tkinter
from datetime import datetime
from tkinter import ttk, messagebox, Canvas
from patientSignUp import *
import openpyxl


class Reception:
    def __init__(self):
        
        # creating the window and placing the icon 
        self.window = tk.Tk()
        self.window.title("Patient Sign-Up")
        self.window.geometry(f"1285x{self.window.winfo_screenheight()-75}+{(self.window.winfo_screenwidth()-1285)//2}+0")
        self.window.configure(bg='#FAF9F6')
        self.window.iconbitmap(r"media/login.ico")
        
        # creating our main frame
        self.main_frame = tk.Frame(self.window, bg='#ffffff')
        
        # creating two other frames inside the main frame
        self.headings_frame = tk.LabelFrame(self.main_frame, bg='#ffffff')
        self.data_view_frame = tk.LabelFrame(self.main_frame,bg='#ffffff')
        
        # creating our labels, entries and buttons
        self.reception_label = tk.Label(self.main_frame,text="Welcome to the reception", bg='#ffffff', fg='#000000', font=("Arial", 30))
        self.date_label = tk.Label(self.headings_frame,text="date : " + datetime.now().strftime('%d/%m/%y'), bg='#FAF9F6', fg='#000000', font=("Arial", 10))
        self.search_label = tk.Label(self.headings_frame,text="Search",bg='#FAF9F6', fg='#000000', font=("Arial", 10))
        self.search_entry = tk.Entry(self.headings_frame,bd=2)
        self.search_button = tk.Button(self.headings_frame,text="Search",command=self.search_func,bg='#00A1D1',fg='#000000',bd=0,relief=tk.GROOVE,width=10)
        self.reset_button = tk.Button(self.headings_frame,text="Reset",command=self.load_data,bg='#00A1D1',fg='#000000',bd=0,relief=tk.GROOVE,width=10)
        self.add_patient_button = tk.Button(self.headings_frame, text="Add Patient", bg='#00A1D1', fg='#000000', font=("Arial", 15), command=self.add_patient,bd=0,relief=tk.GROOVE,width=15)
        self.exit_button = tk.Button(self.headings_frame, text="Exit", bg='#00A1D1', fg='#000000', font=("Arial", 15), command=self.window.destroy,bd=0,relief=tk.GROOVE,width=10)

        # creating our treeview that will show the data of the patients and creating the headers of the columns
        cols = ("Patient Name","Patient ID","Phone Number","Birth Date","Height (cm)","Weight (kg)","Blood Type","Service","Checkup-type","Appointment","Governorate")
        self.tree_view = ttk.Treeview(self.data_view_frame,show="headings",columns=cols,height=self.window.winfo_screenheight()-790)
        self.tree_view.grid(row=0,column=0)
        
        # modifying the width of each column of the treeview
        self.tree_view.column("Patient Name",width=130)
        self.tree_view.column("Phone Number",width=100)
        self.tree_view.column("Patient ID",width=100)
        self.tree_view.column("Birth Date",width=90)
        self.tree_view.column("Height (cm)",width=50)
        self.tree_view.column("Weight (kg)",width=50)
        self.tree_view.column("Blood Type",width=70)
        self.tree_view.column("Governorate",width=90)

        # grid the headings frame with the labels, entries and buttons inside of it
        self.headings_frame.grid(row=1, column=0, sticky="news")
        self.reception_label.grid(row=0, column=0)
        self.date_label.grid(row=1, column=1, padx=50)
        self.search_entry.grid(row=1, column=2, columnspan=2)
        self.search_button.grid(row=1, column=4,padx=5)
        self.reset_button.grid(row=1, column=5, padx=20)
        # removing the reset button after clicking on it
        self.reset_button.grid_remove()
        # putting the add_patient button and the exit button at the far right of the headings frame
        self.headings_frame.columnconfigure(6, weight=1)  
        self.add_patient_button.grid(row=1, column=6, padx=(20, 5), sticky="e")
        self.exit_button.grid(row=1, column=7, padx=(5, 20), sticky="e")
        # grid the data_view_frame
        self.data_view_frame.grid(row=2, column=0)
        # packing our main frame
        self.main_frame.pack(fill="both",expand=1)
        # loading the patients data after 0.1 seconds
        self.window.after(100, self.load_data)
        self.window.mainloop()
        
    # creating a function to load the PatientSignUp page when clicking on the add_patient button
    def add_patient(self):
        self.window.destroy()
        PatientSignUp()

    # loading the patients data from the xlsx file
    def load_data(self):
        path = r"data\patient_data.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        
        list_values = list(sheet.values)
        for col_name in list_values[0]:
            self.tree_view.heading(col_name,text=col_name)
        for value_tuple in list_values[1:]:
            self.tree_view.insert('',tk.END,values=value_tuple)
            
        self.reset_button.grid_remove()
    
    # creating a search function to search for a specific patient using his id
    def search_func(self):
        desired_patient = self.search_entry.get()

        path = r"data\patient_data.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active

        for item in self.tree_view.get_children():
            self.tree_view.delete(item)

        for row in sheet.iter_rows(values_only=True):
            if str(row[1]) == desired_patient:
                self.tree_view.insert("", tk.END, values=row)

        workbook.close()
        
        self.reset_button.grid()
        
    
if __name__ == "__main__":
    app = Reception()
