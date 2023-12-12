from reception_page import *
import reception_page as rc
import openpyxl


class PatientSignUp:
    def __init__(self):
        
        # creating the window and placing the icon 
        self.window = tkinter.Tk()
        self.window.title("Add Patient")
        self.window.geometry(f"{self.window.winfo_screenwidth()-10}x{self.window.winfo_screenheight()-75}+0+0")
        self.window.configure(bg='#ffffff')
        self.window.iconbitmap(r"media/signup.ico")

        # creating our main frame
        self.main_frame = tkinter.Frame(self.window, bg='#ffffff')

        # creating a scroll bar using canvas library
        self.canvas = Canvas(self.main_frame)
        self.canvas.pack(side="left",fill="both",expand=1)
        
        self.scrollbar = ttk.Scrollbar(self.main_frame,orient="vertical",command=self.canvas.yview)
        self.scrollbar.pack(side="right",fill="y")
        
        self.canvas.configure(yscrollcommand=self.scrollbar.set)
        self.canvas.bind('<Configure>',lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        
        self.canvas_frame = tkinter.Frame(self.canvas)
        self.canvas.create_window((0,0),window=self.canvas_frame,anchor="nw")
        
        # creating our labels, entries and buttons
        self.signup_label = tkinter.Label(self.canvas_frame, text="New Patient", bg='#ffffff', fg='#000000', font=("Arial", 30))
        self.patient_label = tkinter.Label(self.canvas_frame, text="Patient Name", bg='#ffffff', fg='#000000', font=("Arial", 16))
        self.patient_entry = tkinter.Entry(self.canvas_frame, font=("Arial", 16))
        self.ID_label = tkinter.Label(self.canvas_frame, text="Patient ID", bg='#ffffff', fg='#000000', font=("Arial", 16))
        self.ID_entry = tkinter.Entry(self.canvas_frame, font=("Arial", 16))
        self.phone_label = tkinter.Label(self.canvas_frame, text="Patient Phone", bg='#ffffff', fg='#000000', font=("Arial", 16))
        self.phone_entry = tkinter.Entry(self.canvas_frame, font=("Arial", 16))
        self.birth_label = tkinter.Label(self.canvas_frame, text="Date of Birth", bg='#ffffff', fg='#000000', font=("Arial", 16))
        self.birth_entry = tkinter.Entry(self.canvas_frame, font=("Arial", 16))
        self.date_note_label = tkinter.Label(self.canvas_frame, text="input in DD/MM/YYYY format", bg='#ffffff', fg='red', font=("Arial", 12))
        self.height_label = tkinter.Label(self.canvas_frame, text="Height", bg='#ffffff', fg='#000000', font=("Arial", 16))
        self.height_entry = ttk.Entry(self.canvas_frame, font=("Arial", 16))
        self.weight_label = tkinter.Label(self.canvas_frame, text="Weight", bg='#ffffff', fg='#000000', font=("Arial", 16))
        self.weight_entry = ttk.Entry(self.canvas_frame, font=("Arial", 16))
        self.bloodtype_label = tkinter.Label(self.canvas_frame, text="Blood Type", bg='#ffffff', fg='#000000', font=("Arial", 16))
        self.bloodtype_combobox = ttk.Combobox(self.canvas_frame, values=["A+", "A-", "B+", "B-", "AB+", "AB-", "O+", "O-"],state="readonly")
        self.service_label = tkinter.Label(self.canvas_frame, text="Service", bg='#ffffff', fg='#000000', font=("Arial", 16))
        self.service_combobox = ttk.Combobox(self.canvas_frame, values=["medical examination" , "pharmacy"],state="readonly")
        self.checkup_label = tkinter.Label(self.canvas_frame, text="checkup type", bg='#ffffff', fg='#000000', font=("Arial", 16))
        self.checkup_combobox = ttk.Combobox(self.canvas_frame, values=["General checkup", "Dermatological checkup", "Nasal examination", "Ear and throat examination", "Cardiovascular checkup", "Ophthalmological checkup", "Dental checkup", "Orthopedic checkup", "Gynecological checkup", "Urological checkup","none"],state="readonly")
        self.checkup_note_label = tkinter.Label(self.canvas_frame, text="Note: Select 'none' if you choose a pharmacy service", bg='#ffffff', fg='red', font=("Arial", 12))
        self.booking_label = tkinter.Label(self.canvas_frame,text="Booking an appointment", bg='#ffffff', fg='#000000', font=("Arial", 16))
        self.booking_date = tkinter.Entry(self.canvas_frame, font=("Arial", 16))
        self.booking_date_note_label = tkinter.Label(self.canvas_frame, text="input in DD/MM/YYYY format", bg='#ffffff', fg='red', font=("Arial", 12))
        self.governorate_label = tkinter.Label(self.canvas_frame, text="governorate", bg='#ffffff', fg='#000000', font=("Arial", 16))
        self.governorate_combobox = ttk.Combobox(self.canvas_frame, values=["Cairo", "Alexandria", "Giza", "Shubra El-Kheima", "Port Said", "Suez", "Luxor", "Mansoura", "El-Mahalla El-Kubra", "Tanta", "Asyut", "Ismailia", "Fayyum", "Zagazig", "Aswan", "Damietta", "Damanhur", "Minya", "Beni Suef", "Hurghada", "Qena", "Sohag", "Banha", "Kafr El-Sheikh", "Arish", "Mallawi"], state="readonly")
        self.signup_button = tkinter.Button(self.canvas_frame, text="Add", bg='#00A1D1', fg='#000000', font=("Arial", 16),command=self.signup,bd=0,relief=tkinter.GROOVE,width=10)
        
        # grid our labels, entries and buttons
        self.signup_label.grid(row=0, column=1, columnspan=2, sticky="news", pady=40)
        self.patient_label.grid(row=1, column=0)
        self.patient_entry.grid(row=1, column=1, pady=20)
        self.ID_label.grid(row=2, column=0)
        self.ID_entry.grid(row=2, column=1, pady=20)
        self.phone_label.grid(row=3, column=0)
        self.phone_entry.grid(row=3, column=1, pady=20)
        self.birth_label.grid(row=4, column=0)
        self.birth_entry.grid(row=4, column=1, pady=20)
        self.date_note_label.grid(row=4,column=2)
        self.height_label.grid(row=5, column=0)
        self.height_entry.grid(row=5, column=1, pady=20)
        self.weight_label.grid(row=6, column=0)
        self.weight_entry.grid(row=6, column=1, pady=20)
        self.bloodtype_label.grid(row=7, column=0)
        self.bloodtype_combobox.grid(row=7, column=1, pady=20)
        self.service_label.grid(row=8, column=0)
        self.service_combobox.grid(row=8, column=1, pady=20)
        self.checkup_label.grid(row=9, column=0)
        self.checkup_combobox.grid(row=9, column=1, pady=20)
        self.checkup_note_label.grid(row=9, column=2, padx=10)
        self.booking_label.grid(row=10, column=0)
        self.booking_date.grid(row=10, column=1, pady=20)
        self.booking_date_note_label.grid(row=10,column=2)
        self.governorate_label.grid(row=11, column=0)
        self.governorate_combobox.grid(row=11, column=1, pady=20)
        self.signup_button.grid(row=12, column=1,columnspan=2, pady=30)
        
        # packing our main frame
        self.main_frame.pack(fill="both",expand=1)
        
        self.window.mainloop()

    # data validation, we check if the entered data is correct following specific criteria, if the data is correct it returns true otherwise it returns false
    def validated_data(self):
        is_validated = True
        user_data = {"patient_name":self.patient_entry.get(),
                   "patient_id":self.ID_entry.get(),
                   "patient_phone":self.phone_entry.get(),
                   "birth_date":self.birth_entry.get(),
                   "height":self.height_entry.get(),
                   "weight":self.weight_entry.get(),
                   "blood_type":self.bloodtype_combobox.get(),
                   "service":self.service_combobox.get(),
                   "checkup-type":self.checkup_combobox.get(),
                   "appointment":self.booking_date.get(),
                   "governate":self.governorate_combobox.get()}
        for key,value in user_data.items():
            if key in ("birth_date", "appointment"):
                continue 
            # if there is an entry missing it shows a warning message with that specific entry
            if not len(value):
                is_validated=False
                messagebox.showerror("warning",f"the {key} is missing")
                return is_validated
        # if the id already exist it shows an error message
        if self.exist_id(self.ID_entry.get()):
            is_validated=False
            messagebox.showerror("warning","the patient id already exists")
        # if the length of the id is not 14 it shows an error message
        if len(self.ID_entry.get()) != 14:
            is_validated=False
            messagebox.showerror("warning","the patient id must be exactly 14 numbers")
        # if the length of the phone number is not 11 it shows an error message
        if len(self.phone_entry.get()) != 11:
            is_validated=False
            messagebox.showerror("warning","the patient phone must be exactly 11 numbers")
        # if the age of the patient is less than 18 it shows an error message
        if self.calculate_age(self.birth_entry.get()) < 18:
            is_validated=False
            messagebox.showerror("warning","you must be 18 or older")
        # if the height of the patient is not valid  it shows an error message
        if int(self.height_entry.get()) > 250 or int(self.height_entry.get()) < 60:
            is_validated=False
            messagebox.showerror("warning","enter a valid height in cm")
        # if the weight of the patient is not valid  it shows an error message
        if int(self.weight_entry.get()) > 600 or int(self.height_entry.get()) < 20:
            is_validated=False
            messagebox.showerror("warning","enter a valid weight in kg")
        else:
            return is_validated

    # a signup function that will execute when clicking on the add button, and it will get the patient data after validating it using the validated_data func and then shows a success message that the patient was added
    def signup(self):
        if self.validated_data():
            patient_name = self.patient_entry.get()
            patient_id = self.ID_entry.get()
            patient_phone = self.phone_entry.get()
            birth_date = self.birth_entry.get()
            height = self.height_entry.get()
            weight = self.weight_entry.get()
            blood_type = self.bloodtype_combobox.get()
            service = self.service_combobox.get()
            checkup = self.checkup_combobox.get()
            appointment = self.booking_date.get()
            governate = self.governorate_combobox.get()

            path = r"data\patient_data.xlsx"
            workbook = openpyxl.load_workbook(path)
            sheet = workbook.active
            row_values = [patient_name,patient_id,patient_phone,birth_date,height,weight,blood_type,service,checkup,appointment,governate]
            sheet.append(row_values)
            workbook.save(path)
            
            messagebox.showinfo(title="signed up",message="Patient added successfully")
            self.window.destroy()
            rc.Reception()

    # calculate the age of the patient to check if he is eligible or not in the validated_data func
    def calculate_age(self, birth_date):
        birth_day, birth_month, birth_year = birth_date.split('/')
        today = datetime.today()
        age = today.year - int(birth_year) - ((today.month, today.day) < (int(birth_month), int(birth_day)))
        return age

    # the function that will be used to check if the entered id already exists or not 
    def exist_id(self, ID):
        ID = self.ID_entry.get()
        
        path = r"data\patient_data.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        patient_found = False
        for row in sheet.iter_rows(values_only=True):
            if row[1] == int(ID.strip()) :
                patient_found = True
                break
        return patient_found

        
if __name__ == "__main__":
    app = PatientSignUp()
